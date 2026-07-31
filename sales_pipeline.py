"""
Automated Reporting Pipeline -- Auto-Adapting Edition
---------------------------------------------------------------------------
Works with ANY .xlsx, .xls, or .csv file, whatever its header names or
layout are. It does NOT rely on a fixed list of expected column names.

How the auto-detection works:
  1. HEADER ROW DETECTION -- scans the first ~15 rows of each file to find
     which row is actually the header (skips title rows, blank rows, merged
     banner cells that real-world files often have above the real table).
  2. COLUMN TYPE DETECTION -- for every column, decides if it's:
       - a DATE column   (parses as a date for most of its values)
       - a NUMBER column (parses as numeric for most of its values --
         handles "1,234", "₹1,234.50", "1234" etc.)
       - a CATEGORY/label column (everything else -- names, regions,
         products, dealer names, etc.)
  3. COMBINE -- every cleaned file is stacked into one master table, with
     a "Source File" column so you can always trace a row back to its
     original file.
  4. REPORT -- builds a formula-driven Excel report:
       - Raw Data sheet: the full combined, cleaned table
       - Summary sheet: auto-picks the best category column to group by,
         and SUMIFS-totals every detected number column against it,
         with KPI cards and a chart -- all live formulas, not hardcoded.

Required environment variables (set as Render/GitHub secrets):
  GOOGLE_SERVICE_ACCOUNT_JSON   full JSON key content of your service account
  RAW_FOLDER_ID                 Drive folder ID with the manager-uploaded files
  OUTPUT_FOLDER_ID              Drive folder ID where the report should be saved
"""
import os
import re
import glob
import tempfile
import logging
from datetime import datetime


def clean_labels(dl, **overrides):
    """
    Force every DataLabelList flag to an explicit True/False instead of
    leaving it as None. openpyxl leaves unset flags (showCatName,
    showSerName, showLegendKey, showPercent, showBubbleSize, showVal) as
    None, meaning they simply aren't written to the chart XML at all.
    Google Sheets happens to treat "not present" as False, so charts
    opened there look fine -- but Excel on desktop treats several of
    these as True by default when they're missing, so labels come out as
    "Series, Category, Value" all concatenated together instead of just
    the single value you wanted. Setting every flag explicitly makes
    both apps render the same way.
    """
    dl.showLegendKey = False
    dl.showVal = False
    dl.showCatName = False
    dl.showSerName = False
    dl.showPercent = False
    dl.showBubbleSize = False
    for k, v in overrides.items():
        setattr(dl, k, v)
    return dl

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, LineChart, DoughnutChart, ScatterChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

from gdrive_utils import (
    get_drive_service, list_data_files_in_folder, download_file, upload_or_replace_file,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("pipeline")

REPORT_FILENAME = "Master_Report.xlsx"

# ---------------------------------------------------------------------------
# STEP 1: SMART FILE READING -- auto-detects the real header row
# ---------------------------------------------------------------------------

def smart_parse_dates(series):
    """
    Parses a column to dates without the month/day-swap bug that a blanket
    dayfirst=True causes on ISO-style strings. Unambiguous ISO dates
    (YYYY-MM-DD) are parsed as-is; anything else falls back to day-first
    parsing, since that's the convention most Indian business files use.
    """
    if pd.api.types.is_datetime64_any_dtype(series):
        return series
    s = series.astype(str)
    iso_mask = s.str.match(r"^\d{4}-\d{1,2}-\d{1,2}")
    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    if iso_mask.any():
        parsed.loc[iso_mask] = pd.to_datetime(series[iso_mask], errors="coerce")
    if (~iso_mask).any():
        parsed.loc[~iso_mask] = pd.to_datetime(series[~iso_mask], errors="coerce", dayfirst=True)
    return parsed


def _looks_numeric(value):
    if value is None:
        return False
    s = str(value).strip().replace(",", "").replace("₹", "").replace("$", "")
    return bool(re.match(r"^-?\d+(\.\d+)?%?$", s))


def detect_header_row(raw_df, max_scan=15):
    """Scores each of the first `max_scan` rows on how header-like it looks,
    and how data-like the row right after it looks, then returns the best."""
    best_row, best_score = 0, -1e9
    limit = min(max_scan, len(raw_df))
    for i in range(limit):
        row = raw_df.iloc[i]
        filled = row.notna().sum()
        if filled == 0:
            continue
        text_like = sum(
            1 for v in row if isinstance(v, str) and v.strip() and not _looks_numeric(v)
        )
        score = text_like - (row.isna().sum() * 0.5)
        if i + 1 < len(raw_df):
            next_row = raw_df.iloc[i + 1]
            data_like = sum(1 for v in next_row if _looks_numeric(v) or v is not None)
            score += data_like * 0.2
        if score > best_score:
            best_score = score
            best_row = i
    return best_row


def smart_read(filepath):
    """Reads a csv/xls/xlsx file, auto-detecting which row is the real header."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".csv":
        raw = pd.read_csv(filepath, header=None, dtype=object)
        header_row = detect_header_row(raw)
        df = pd.read_csv(filepath, header=header_row, dtype=object)
    else:
        raw = pd.read_excel(filepath, header=None, dtype=object)
        header_row = detect_header_row(raw)
        df = pd.read_excel(filepath, header=header_row, dtype=object)

    df.columns = [str(c).strip() if not str(c).startswith("Unnamed") else f"Column_{i+1}"
                  for i, c in enumerate(df.columns)]
    # drop fully blank rows/columns that sneak in from title/footer bands
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    return df


# ---------------------------------------------------------------------------
# STEP 2: COLUMN TYPE DETECTION
# ---------------------------------------------------------------------------

def classify_columns(df):
    """Returns (date_col, numeric_cols, category_cols) for a dataframe,
    without assuming any particular column names."""
    date_col = None
    numeric_cols = []
    category_cols = []

    # Column-name patterns that mean "this is an identifier/code, not a
    # measure" are handled by the module-level ID_PATTERN.

    # pass 1: prefer a column whose name hints it's a date
    for col in df.columns:
        name_lower = str(col).lower()
        if date_col is None and any(k in name_lower for k in ["date", "month", "period"]):
            parsed = smart_parse_dates(df[col])
            if parsed.notna().mean() > 0.4:
                date_col = col

    for col in df.columns:
        if col == date_col:
            continue
        name_lower = str(col).lower()
        series = df[col]

        # Identifier/code columns are always treated as labels, never as
        # a summable metric, regardless of how "numeric" their values look.
        if ID_PATTERN.search(name_lower):
            category_cols.append(col)
            continue

        cleaned = series.astype(str).str.replace(",", "", regex=False)\
                                     .str.replace("₹", "", regex=False)\
                                     .str.replace("$", "", regex=False)\
                                     .str.replace("%", "", regex=False).str.strip()
        numeric = pd.to_numeric(cleaned, errors="coerce")
        if numeric.notna().mean() > 0.6:
            # Note: we deliberately do NOT use "almost every value is
            # unique" as a signal for excluding a column here. Genuine
            # continuous financial amounts (revenue, invoice totals, etc.)
            # very often have high uniqueness too, especially with more
            # rows -- that heuristic caused real amount columns to be
            # mistaken for identifiers. Name-based ID_PATTERN detection
            # above is the reliable signal; if a numeric column's name
            # doesn't match it, we trust it as a genuine metric.
            numeric_cols.append(col)
        elif date_col is None:
            parsed = smart_parse_dates(series)
            if parsed.notna().mean() > 0.6:
                date_col = col
            else:
                category_cols.append(col)
        else:
            category_cols.append(col)

    return date_col, numeric_cols, category_cols


# Column-name patterns that mean "this is an identifier/code, not a
# measure or a friendly grouping label" -- e.g. "Location Code", "GST No",
# "Invoice No" -- these should never be summed, nor chosen as the
# category to group charts by, even if they happen to contain a
# location/name-ish word as part of a compound name.
ID_PATTERN = re.compile(
    r"\b(code|no\.?|number|id|gstin|gst\s*no|pan|pincode|pin\s*code|"
    r"zip|phone|mobile|contact|aadhar|account\s*no|invoice\s*no|"
    r"reference|ref\s*no|serial)\b",
    re.IGNORECASE,
)


def score_numeric_column(name):
    """
    Ranks how likely a numeric column is to be a headline business metric
    worth featuring on KPI cards / top of the summary, vs. a secondary or
    derived figure (a tax break-up line, a day-count, an allowance, etc).
    Higher score = more important. Uses whole-word matching so short
    keywords like 'ta' or 'day' don't false-match inside 'Target'/'Total'.
    """
    name_lower = str(name).lower()
    HIGH_KEYWORDS = ["revenue", "sales", "amount", "amt", "total", "payment",
                      "received", "collection", "target", "pending", "gst",
                      "due", "balance", "net", "gross", "invoice", "value",
                      "cost", "price", "profit"]
    LOW_KEYWORDS = ["cess", "surcharge", "tds", "credit", "day", "ta"]
    score = 0
    for k in HIGH_KEYWORDS:
        if re.search(rf"\b{re.escape(k)}\b", name_lower):
            score += 2
    for k in LOW_KEYWORDS:
        if re.search(rf"\b{re.escape(k)}\b", name_lower):
            score -= 1
    if "%" in name_lower:
        score -= 1
    return score


def find_status_column(category_cols):
    """Returns a column that looks like a status/outcome flag (e.g. 'Paid',
    'Pending', 'Short Payment') if one exists -- these make an excellent
    donut/breakdown chart for billing & collections style data."""
    for col in category_cols:
        if re.search(r"\bstatus\b", str(col).lower()):
            return col
    return None


def pick_secondary_column(category_cols, group_col, df, n_rows):
    """Picks a second, genuinely different breakdown dimension for the
    Top-N chart -- excludes the primary group_col and anything that's
    really a date in disguise (e.g. a 'Month' column that ended up as a
    category because a different date column was already claimed)."""
    candidates = [c for c in category_cols if c != group_col]
    non_date_candidates = []
    for c in candidates:
        parsed = smart_parse_dates(df[c])
        if parsed.notna().mean() > 0.5:
            continue
        non_date_candidates.append(c)
    if not non_date_candidates:
        return None
    return pick_grouping_column(non_date_candidates, df, n_rows)


def pick_grouping_column(category_cols, df, n_rows):
    """Picks the best column to group the summary by: prefers an obviously
    label-like name, otherwise the category column with a sensible number
    of distinct values (not near-unique, not a single repeated value).
    Identifier/code columns (Location Code, Invoice No, etc.) are excluded
    even if they contain a location/name-ish word as part of a compound
    name -- they're not friendly, human-readable groupings."""
    preferred_keywords = ["region", "product", "item", "name", "party", "dealer",
                          "branch", "category", "zone", "state", "city", "sku",
                          "district", "particulars", "client", "vendor",
                          "location", "office", "department", "customer",
                          "hospital", "college", "company", "account"]
    candidates = [c for c in category_cols if not ID_PATTERN.search(str(c).lower())]
    if not candidates:
        candidates = category_cols  # fall back rather than returning nothing

    for col in candidates:
        if any(k in str(col).lower() for k in preferred_keywords):
            return col
    best_col, best_score = None, -1
    for col in candidates:
        n_unique = df[col].nunique(dropna=True)
        if 1 < n_unique <= max(50, n_rows * 0.5):
            score = -abs(n_unique - 8)  # sweet spot around ~8 groups
            if score > best_score:
                best_score = score
                best_col = col
    return best_col or (candidates[0] if candidates else None)


# ---------------------------------------------------------------------------
# STEP 3: CLEAN + COMBINE
# ---------------------------------------------------------------------------

def clean_one_file(filepath):
    df = smart_read(filepath)
    date_col, numeric_cols, category_cols = classify_columns(df)

    for c in numeric_cols:
        cleaned = df[c].astype(str).str.replace(",", "", regex=False)\
                                    .str.replace("₹", "", regex=False)\
                                    .str.replace("$", "", regex=False)\
                                    .str.replace("%", "", regex=False).str.strip()
        df[c] = pd.to_numeric(cleaned, errors="coerce")
    if date_col:
        df[date_col] = smart_parse_dates(df[date_col])

    df["Source File"] = os.path.basename(filepath)
    df.attrs["date_col"] = date_col
    df.attrs["numeric_cols"] = numeric_cols
    df.attrs["category_cols"] = category_cols
    return df


def is_useful_numeric_column(series, min_fill_ratio=0.3):
    """
    Rejects numeric columns that aren't actually useful to chart/sum:
    mostly blank (a sparse tax break-up line that's rarely filled in),
    entirely zero, or constant (every row has the same value -- no
    variation to show on a chart).
    """
    total = len(series)
    if total == 0:
        return False
    non_null = series.notna().sum()
    if (non_null / total) < min_fill_ratio:
        return False
    filled = series.fillna(0)
    if (filled != 0).sum() == 0:
        return False
    if series.nunique(dropna=True) <= 1:
        return False
    return True


def is_useful_category_column(series, min_fill_ratio=0.3):
    """Rejects category columns that are mostly blank or have no real
    variation (every row the same label -- not useful for grouping)."""
    total = len(series)
    if total == 0:
        return False
    non_null = series.notna().sum()
    if (non_null / total) < min_fill_ratio:
        return False
    if series.nunique(dropna=True) <= 1:
        return False
    return True


def build_master_dataset(local_dir):
    files = (glob.glob(os.path.join(local_dir, "*.xlsx"))
             + glob.glob(os.path.join(local_dir, "*.xls"))
             + glob.glob(os.path.join(local_dir, "*.csv")))
    if not files:
        raise RuntimeError("No .xlsx/.xls/.csv files found in the Drive raw folder.")

    frames = []
    all_numeric, all_category = [], []
    common_date_col = None

    for f in files:
        try:
            df = clean_one_file(f)
            frames.append(df)
            for c in df.attrs["numeric_cols"]:
                if c not in all_numeric:
                    all_numeric.append(c)
            for c in df.attrs["category_cols"]:
                if c not in all_category:
                    all_category.append(c)
            if not common_date_col and df.attrs["date_col"]:
                common_date_col = df.attrs["date_col"]
            log.info(f"Cleaned {os.path.basename(f)}: {len(df)} rows, "
                      f"date_col={df.attrs['date_col']}, numeric={df.attrs['numeric_cols']}")
        except Exception as e:
            log.warning(f"Skipped {os.path.basename(f)} due to error: {e}")

    if not frames:
        raise RuntimeError("Every file failed cleaning -- check file format.")

    master = pd.concat(frames, ignore_index=True, sort=False)

    useful_numeric = [c for c in all_numeric if is_useful_numeric_column(master[c])]
    dropped_numeric = [c for c in all_numeric if c not in useful_numeric]
    if dropped_numeric:
        log.info(f"Dropping low-quality numeric columns (mostly blank/zero/constant): {dropped_numeric}")

    useful_category = [c for c in all_category if is_useful_category_column(master[c])]
    dropped_category = [c for c in all_category if c not in useful_category]
    if dropped_category:
        log.info(f"Dropping low-quality category columns (mostly blank/no variation): {dropped_category}")

    useful_numeric.sort(key=score_numeric_column, reverse=True)
    return master, useful_numeric, useful_category, common_date_col


# ---------------------------------------------------------------------------
# STEP 4: REPORT BUILDING (openpyxl, formula-driven)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E78")
LABEL_FONT = Font(name="Arial", bold=True, size=11)
NORMAL_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_raw_sheet(wb, master, date_col):
    ws = wb.active
    ws.title = "Raw Data"
    headers = list(master.columns)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(master.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, 1):
            col_name = headers[col_idx - 1]
            if isinstance(value, float) and pd.isna(value):
                value = None
            if pd.isna(value):
                value = None
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = NORMAL_FONT
            c.border = BORDER
            if col_name == date_col and value is not None:
                c.number_format = "DD-MMM-YYYY"

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "A2"
    return len(master) + 1, {h: i + 1 for i, h in enumerate(headers)}


def write_summary_sheet(wb, master, last_raw_row, col_index, numeric_cols, group_col):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Auto-generated {datetime.now().strftime('%d-%b-%Y %H:%M')} from {last_raw_row - 1} rows"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

    # KPI cards: total of each numeric column
    kpi_row = 4
    col_cursor = 1
    for metric in numeric_cols[:4]:
        letter = get_column_letter(col_index[metric])
        label_cell = ws.cell(row=kpi_row, column=col_cursor, value=f"Total {metric}")
        label_cell.font = LABEL_FONT
        value_cell = ws.cell(row=kpi_row, column=col_cursor + 1,
                              value=f"=SUM('Raw Data'!{letter}2:{letter}{last_raw_row})")
        value_cell.number_format = "#,##0"
        value_cell.font = Font(name="Arial", bold=True, size=14, color="1F4E78")
        col_cursor += 3

    if not group_col or not numeric_cols:
        ws["A6"] = "No suitable category column or numeric metric detected for a breakdown table."
        return

    groups = sorted([g for g in master[group_col].dropna().unique()])
    group_letter = get_column_letter(col_index[group_col])

    start_row = 7
    ws.cell(row=start_row, column=1, value=group_col).font = HEADER_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    for col_idx, metric in enumerate(numeric_cols, start=2):
        c = ws.cell(row=start_row, column=col_idx, value=metric)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx, g in enumerate(groups, start=start_row + 1):
        ws.cell(row=r_idx, column=1, value=g).font = LABEL_FONT
        for c_idx, metric in enumerate(numeric_cols, start=2):
            metric_letter = get_column_letter(col_index[metric])
            formula = (
                f"=SUMIFS('Raw Data'!${metric_letter}$2:${metric_letter}${last_raw_row},"
                f"'Raw Data'!${group_letter}$2:${group_letter}${last_raw_row},$A{r_idx})"
            )
            cell = ws.cell(row=r_idx, column=c_idx, value=formula)
            cell.number_format = "#,##0"
            cell.border = BORDER

    total_row = start_row + len(groups) + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(name="Arial", bold=True)
    for c_idx in range(2, len(numeric_cols) + 2):
        col_letter = get_column_letter(c_idx)
        cell = ws.cell(row=total_row, column=c_idx,
                        value=f"=SUM({col_letter}{start_row+1}:{col_letter}{total_row-1})")
        cell.number_format = "#,##0"
        cell.font = Font(name="Arial", bold=True)

    if numeric_cols:
        data_range = f"B{start_row+1}:{get_column_letter(len(numeric_cols)+1)}{total_row-1}"
        median_est = master[numeric_cols[0]].median()
        threshold = median_est * 0.5 if pd.notna(median_est) and median_est > 0 else 0
        if threshold > 0:
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(operator="lessThan", formula=[str(threshold)],
                           fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
            )

        chart = BarChart()
        chart.title = f"{numeric_cols[0]} by {group_col}"
        chart.y_axis.title = numeric_cols[0]
        chart.x_axis.title = group_col
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(groups))
        cats_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(groups))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width, chart.height = 16, 8
        ws.add_chart(chart, f"A{total_row + 3}")

    for col_idx in range(1, len(numeric_cols) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def write_dashboard_sheet(wb, master, last_raw_row, col_index, numeric_cols, group_col, date_col, category_cols, status_col=None):
    """
    Builds a compact, dark-theme 'Dashboard' sheet with GENUINE multi-filter
    cross-filtering, Power-BI style, using only native Excel formulas (no
    macros -- openpyxl can't create real PivotTables/Slicers, so this is
    the closest working equivalent):

      - Up to 3 dropdown filters: by <group_col>, by <status_col> (if
        found), and by Month (if a date column was found).
      - EVERY chart's underlying values are SUMPRODUCT formulas that
        respect whichever filters are currently active, EXCEPT a chart
        never filters by its own breakdown dimension (e.g. the "by
        Group" chart still shows all groups even if a Group filter is
        set, so you can still compare -- standard BI convention). Cross-
        filtering from a *different* dimension (e.g. filtering by Status
        while looking at the "by Group" chart) works on every chart.
      - Chart categories (the list of groups/statuses/dates shown) stay
        fixed so the charts don't reshape oddly; only the values move.
    """
    ws = wb.create_sheet("Dashboard")

    PAGE_BG = "0B1F3A"
    BANNER_BG = "081530"
    CARD_COLORS = ["2E75B6", "2E8B57", "7030A0", "C55A11", "1F9C9C"]
    PANEL_BG = "13294B"
    SECTION_BG = "13294B"
    WHITE = "FFFFFF"
    LIGHT_GREY = "C9D6E8"

    last_col = 24
    last_row = 82

    dark_fill = PatternFill(start_color=PAGE_BG, end_color=PAGE_BG, fill_type="solid")
    for r in range(1, last_row):
        for c in range(1, last_col):
            ws.cell(row=r, column=c).fill = dark_fill
    for c in range(1, last_col):
        ws.column_dimensions[get_column_letter(c)].width = 11
    for r in range(1, last_row):
        ws.row_dimensions[r].height = 16

    # ---- title banner ----
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=last_col - 1)
    banner = ws.cell(row=1, column=1, value="DASHBOARD")
    banner.font = Font(name="Arial", bold=True, size=20, color=WHITE)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    for r in (1, 2):
        for c in range(1, last_col):
            ws.cell(row=r, column=c).fill = PatternFill(start_color=BANNER_BG, end_color=BANNER_BG, fill_type="solid")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    if not numeric_cols:
        ws.cell(row=4, column=1, value="No numeric metrics detected.").font = Font(color=WHITE)
        return

    groups = sorted([g for g in master[group_col].dropna().unique()]) if group_col else []
    statuses = sorted([s for s in master[status_col].dropna().unique()]) if status_col else []
    months = []
    if date_col and master[date_col].notna().any():
        periods = sorted(master[date_col].dropna().dt.to_period("M").unique())
        months = [p.strftime("%b-%Y") for p in periods]

    raw_group_letter = get_column_letter(col_index[group_col]) if group_col else None
    raw_status_letter = get_column_letter(col_index[status_col]) if status_col else None
    raw_date_letter = get_column_letter(col_index[date_col]) if date_col else None

    # ==================================================================
    # FILTER PANEL -- up to 3 dropdowns, each backed by a hidden list
    # ==================================================================
    filters = []
    list_col = 26
    frow = 6
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)
    hdr = ws.cell(row=5, column=1, value="FILTERS")
    hdr.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    for c in range(1, 4):
        ws.cell(row=5, column=c).fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")

    def add_filter(label, values, kind, raw_letter, key):
        nonlocal frow, list_col
        if not values:
            return
        cell_ref = f"$B${frow}"
        ws.cell(row=frow, column=1, value=f"{label}:").font = Font(name="Arial", size=9, color=LIGHT_GREY)
        fcell = ws.cell(row=frow, column=2, value="All")
        fcell.font = Font(name="Arial", bold=True, size=10, color="FFD966")
        fcell.fill = PatternFill(start_color=PANEL_BG, end_color=PANEL_BG, fill_type="solid")
        fcell.border = BORDER

        ws.cell(row=3, column=list_col, value="All")
        for i, v in enumerate(values, start=4):
            ws.cell(row=i, column=list_col, value=v)
        list_letter = get_column_letter(list_col)
        dv = DataValidation(type="list",
                             formula1=f"=${list_letter}$3:${list_letter}${3 + len(values)}",
                             allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(fcell)
        ws.column_dimensions[list_letter].hidden = True
        list_col += 1
        filters.append({"key": key, "cell": cell_ref, "kind": kind, "letter": raw_letter})
        frow += 1

    add_filter(group_col, groups, "exact", raw_group_letter, group_col)
    add_filter(status_col, statuses, "exact", raw_status_letter, status_col)
    add_filter("Month", months, "month", raw_date_letter, date_col)

    ws.cell(row=frow, column=1, value="(dropdown arrow ->)").font = Font(name="Arial", italic=True, size=7, color=LIGHT_GREY)

    # ==================================================================
    # KPI CARDS -- reflect ALL active filters combined
    # ==================================================================
    def cond_expr(f):
        rng = f"'Raw Data'!${f['letter']}$2:${f['letter']}${last_raw_row}"
        if f["kind"] == "month":
            return f'(TEXT({rng},"mmm-yyyy")=IF({f["cell"]}="All",TEXT({rng},"mmm-yyyy"),{f["cell"]}))'
        return f'({rng}=IF({f["cell"]}="All",{rng},{f["cell"]}))'

    def metric_formula(metric, exclude_keys=(), extra_cond=None):
        metric_letter = get_column_letter(col_index[metric])
        metric_rng = f"'Raw Data'!${metric_letter}$2:${metric_letter}${last_raw_row}"
        parts = []
        if extra_cond:
            parts.append(extra_cond)
        parts += [cond_expr(f) for f in filters if f["key"] not in exclude_keys]
        if parts:
            return "=SUMPRODUCT(" + "*".join(f"({p})" for p in parts) + f"*({metric_rng}))"
        return f"=SUM({metric_rng})"

    kpi_row = 5
    card_width = 4
    card_start_col = 5
    for i, metric in enumerate(numeric_cols[:4]):
        col_start = card_start_col + i * card_width
        color = CARD_COLORS[i % len(CARD_COLORS)]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        col_end = col_start + card_width - 2
        ws.merge_cells(start_row=kpi_row, start_column=col_start, end_row=kpi_row, end_column=col_end)
        ws.merge_cells(start_row=kpi_row + 1, start_column=col_start, end_row=kpi_row + 2, end_column=col_end)
        for r in range(kpi_row, kpi_row + 3):
            for c in range(col_start, col_start + card_width - 1):
                ws.cell(row=r, column=c).fill = fill

        label_cell = ws.cell(row=kpi_row, column=col_start, value=metric.upper())
        label_cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        val_cell = ws.cell(row=kpi_row + 1, column=col_start, value=metric_formula(metric))
        val_cell.font = Font(name="Arial", bold=True, size=16, color=WHITE)
        val_cell.number_format = "#,##0"
        val_cell.alignment = Alignment(horizontal="left", vertical="center")

    def section_header(row, col_start, col_end, text):
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start, value=text.upper())
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = PatternFill(start_color=SECTION_BG, end_color=SECTION_BG, fill_type="solid")

    # ==================================================================
    # LIVE BREAKDOWN TABLES -- every chart's source data is now formulas
    # ==================================================================
    CAT_COL = 30
    STATUS_COL_H = CAT_COL + len(numeric_cols) + 6
    TOPN_COL = STATUS_COL_H + 8
    DATE_COL_H = TOPN_COL + 8
    PCT_COL = DATE_COL_H + 8

    cat_start_row = None
    chart_groups = []
    if group_col and groups:
        group_totals = master.groupby(group_col)[numeric_cols[0]].sum(numeric_only=True)
        chart_groups = group_totals[group_totals != 0].sort_values(ascending=False).index.tolist()[:15]
    if group_col and chart_groups:
        cat_start_row = 3
        ws.cell(row=cat_start_row, column=CAT_COL, value=group_col)
        for i, m in enumerate(numeric_cols):
            ws.cell(row=cat_start_row, column=CAT_COL + 1 + i, value=m)
        for r_idx, g in enumerate(chart_groups, start=cat_start_row + 1):
            ws.cell(row=r_idx, column=CAT_COL, value=g)
            extra = f"('Raw Data'!${raw_group_letter}$2:${raw_group_letter}${last_raw_row}=${get_column_letter(CAT_COL)}{r_idx})"
            for i, m in enumerate(numeric_cols):
                formula = metric_formula(m, exclude_keys={group_col}, extra_cond=extra)
                ws.cell(row=r_idx, column=CAT_COL + 1 + i, value=formula)
        cat_end_row = cat_start_row + len(chart_groups)
        for c in range(CAT_COL, CAT_COL + len(numeric_cols) + 2):
            ws.column_dimensions[get_column_letter(c)].hidden = True

        BAR_PCT_COL = CAT_COL + len(numeric_cols) + 3
        val_letter = get_column_letter(CAT_COL + 1)
        grand_total_cell = f"${get_column_letter(BAR_PCT_COL)}$2"
        ws.cell(row=2, column=BAR_PCT_COL, value=metric_formula(numeric_cols[0], exclude_keys=set()))
        for r in range(cat_start_row + 1, cat_end_row + 1):
            ws.cell(row=r, column=BAR_PCT_COL,
                    value=f"={val_letter}{r}/{grand_total_cell}")
            ws.cell(row=r, column=BAR_PCT_COL).number_format = "0%"
        ws.column_dimensions[get_column_letter(BAR_PCT_COL)].hidden = True

    status_start_row = None
    chart_statuses = []
    if status_col and statuses:
        status_totals = master.groupby(status_col)[numeric_cols[0]].sum(numeric_only=True)
        chart_statuses = status_totals[status_totals != 0].sort_values(ascending=False).index.tolist()
    if status_col and chart_statuses:
        status_start_row = 3
        ws.cell(row=status_start_row, column=STATUS_COL_H, value=status_col)
        ws.cell(row=status_start_row, column=STATUS_COL_H + 1, value=numeric_cols[0])
        for r_idx, s in enumerate(chart_statuses, start=status_start_row + 1):
            ws.cell(row=r_idx, column=STATUS_COL_H, value=s)
            extra = f"('Raw Data'!${raw_status_letter}$2:${raw_status_letter}${last_raw_row}=${get_column_letter(STATUS_COL_H)}{r_idx})"
            formula = metric_formula(numeric_cols[0], exclude_keys={status_col}, extra_cond=extra)
            ws.cell(row=r_idx, column=STATUS_COL_H + 1, value=formula)
        status_end_row = status_start_row + len(chart_statuses)
        ws.column_dimensions[get_column_letter(STATUS_COL_H)].hidden = True
        ws.column_dimensions[get_column_letter(STATUS_COL_H + 1)].hidden = True

    secondary_col = pick_secondary_column(category_cols, group_col, master, len(master))
    top_start_row = None
    if secondary_col:
        sec_totals = master.groupby(secondary_col)[numeric_cols[0]].sum(numeric_only=True)
        top_items = sec_totals[sec_totals != 0].sort_values(ascending=False).head(10).index.tolist()
        sec_letter = get_column_letter(col_index[secondary_col])
        top_start_row = 3
        ws.cell(row=top_start_row, column=TOPN_COL, value=secondary_col)
        ws.cell(row=top_start_row, column=TOPN_COL + 1, value=numeric_cols[0])
        for r_idx, item in enumerate(top_items, start=top_start_row + 1):
            ws.cell(row=r_idx, column=TOPN_COL, value=item)
            extra = f"('Raw Data'!${sec_letter}$2:${sec_letter}${last_raw_row}=${get_column_letter(TOPN_COL)}{r_idx})"
            formula = metric_formula(numeric_cols[0], exclude_keys=set(), extra_cond=extra)
            ws.cell(row=r_idx, column=TOPN_COL + 1, value=formula)
        top_end_row = top_start_row + len(top_items) if top_items else top_start_row
        ws.column_dimensions[get_column_letter(TOPN_COL)].hidden = True
        ws.column_dimensions[get_column_letter(TOPN_COL + 1)].hidden = True
        TOPN_PCT_COL = TOPN_COL + 3
        if top_items:
            val_letter2 = get_column_letter(TOPN_COL + 1)
            topn_grand_total_cell = f"${get_column_letter(TOPN_PCT_COL)}$2"
            ws.cell(row=2, column=TOPN_PCT_COL, value=metric_formula(numeric_cols[0], exclude_keys=set()))
            for r in range(top_start_row + 1, top_end_row + 1):
                ws.cell(row=r, column=TOPN_PCT_COL,
                        value=f"={val_letter2}{r}/{topn_grand_total_cell}")
                ws.cell(row=r, column=TOPN_PCT_COL).number_format = "0%"
            ws.column_dimensions[get_column_letter(TOPN_PCT_COL)].hidden = True
        if not top_items:
            top_start_row = None

    date_start_row = None
    if date_col and master[date_col].notna().any():
        date_totals = master.groupby(date_col)[numeric_cols[0]].sum(numeric_only=True)
        all_dates = date_totals[date_totals != 0].sort_index().index.tolist()
        if len(all_dates) > 1:
            date_start_row = 3
            ws.cell(row=date_start_row, column=DATE_COL_H, value=date_col)
            ws.cell(row=date_start_row, column=DATE_COL_H + 1, value=numeric_cols[0])
            for r_idx, dval in enumerate(all_dates, start=date_start_row + 1):
                dcell = ws.cell(row=r_idx, column=DATE_COL_H, value=pd.Timestamp(dval).to_pydatetime())
                dcell.number_format = "DD-MMM"
                extra = f"('Raw Data'!${raw_date_letter}$2:${raw_date_letter}${last_raw_row}=${get_column_letter(DATE_COL_H)}{r_idx})"
                formula = metric_formula(numeric_cols[0], exclude_keys=set(), extra_cond=extra)
                ws.cell(row=r_idx, column=DATE_COL_H + 1, value=formula)
            date_end_row = date_start_row + len(all_dates)
            ws.column_dimensions[get_column_letter(DATE_COL_H)].hidden = True
            ws.column_dimensions[get_column_letter(DATE_COL_H + 1)].hidden = True

    # ==================================================================
    # ROW 1: donut (status if available, else group share) + bar (by group)
    # ==================================================================
    row1_anchor = 13
    if group_col and cat_start_row:
        cats_ref = Reference(ws, min_col=CAT_COL, min_row=cat_start_row + 1, max_row=cat_end_row)

        if status_col and status_start_row:
            section_header(row1_anchor - 1, 1, 8, f"{numeric_cols[0]} by {status_col} (cross-filtered)")
            donut_cats_ref = Reference(ws, min_col=STATUS_COL_H, min_row=status_start_row + 1, max_row=status_end_row)
            donut_data_ref = Reference(ws, min_col=STATUS_COL_H + 1, min_row=status_start_row, max_row=status_end_row)
        else:
            section_header(row1_anchor - 1, 1, 8, f"{numeric_cols[0]} Share by {group_col}")
            donut_cats_ref = cats_ref
            donut_data_ref = Reference(ws, min_col=BAR_PCT_COL, min_row=cat_start_row, max_row=cat_end_row)

        donut = DoughnutChart()
        donut.visible_cells_only = False
        donut.add_data(donut_data_ref, titles_from_data=True)
        donut.set_categories(donut_cats_ref)
        donut.width, donut.height = 15, 8
        donut.legend.position = "r"
        donut_n = len(chart_statuses) if (status_col and status_start_row) else len(chart_groups)
        if donut_n <= 8:
            donut.dataLabels = DataLabelList()
            if status_col and status_start_row:
                clean_labels(donut.dataLabels, showPercent=True)
            else:
                clean_labels(donut.dataLabels, showVal=True)
                donut.dataLabels.numFmt = "0%"
        ws.add_chart(donut, f"A{row1_anchor}")

        section_header(row1_anchor - 1, 10, 17, f"{numeric_cols[0]} % by {group_col} (cross-filtered)")
        bar = BarChart()
        bar.type = "bar"
        bar.visible_cells_only = False
        bar_data_ref = Reference(ws, min_col=BAR_PCT_COL, min_row=cat_start_row, max_row=cat_end_row)
        bar.add_data(bar_data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.x_axis.numFmt = "0%"
        bar.width, bar.height = 15, 8
        bar.legend = None
        if len(chart_groups) <= 8:
            bar.dataLabels = DataLabelList()
            clean_labels(bar.dataLabels, showVal=True)
            bar.dataLabels.numFmt = "0%"
        ws.add_chart(bar, f"J{row1_anchor}")

    # ==================================================================
    # ROW 2: Top-N (secondary dimension) + Trend -- both cross-filtered
    # ==================================================================
    row2_anchor = row1_anchor + 18
    if secondary_col and top_start_row:
        section_header(row2_anchor - 1, 1, 8, f"Top {secondary_col} % (cross-filtered)")
        top_cats_ref = Reference(ws, min_col=TOPN_COL, min_row=top_start_row + 1, max_row=top_end_row)
        top_bar = BarChart()
        top_bar.type = "bar"
        top_bar.visible_cells_only = False
        top_data_ref = Reference(ws, min_col=TOPN_PCT_COL, min_row=top_start_row, max_row=top_end_row)
        top_bar.add_data(top_data_ref, titles_from_data=True)
        top_bar.set_categories(top_cats_ref)
        top_bar.x_axis.numFmt = "0%"
        top_bar.width, top_bar.height = 15, 8
        top_bar.legend = None
        if len(top_items) <= 8:
            top_bar.dataLabels = DataLabelList()
            clean_labels(top_bar.dataLabels, showVal=True)
            top_bar.dataLabels.numFmt = "0%"
        ws.add_chart(top_bar, f"A{row2_anchor}")

    if date_col and date_start_row:
        section_header(row2_anchor - 1, 10, 17, f"{numeric_cols[0]} Trend (cross-filtered)")
        line = LineChart()
        line.visible_cells_only = False
        line_data_ref = Reference(ws, min_col=DATE_COL_H + 1, min_row=date_start_row, max_row=date_end_row)
        line_cats_ref = Reference(ws, min_col=DATE_COL_H, min_row=date_start_row + 1, max_row=date_end_row)
        line.add_data(line_data_ref, titles_from_data=True)
        line.set_categories(line_cats_ref)
        for s in line.series:
            s.smooth = False
        line.width, line.height = 15, 8
        line.legend = None
        if (date_end_row - date_start_row) <= 8:
            line.dataLabels = DataLabelList()
            clean_labels(line.dataLabels, showVal=True)
            line.dataLabels.numFmt = "#,##0"
        ws.add_chart(line, f"J{row2_anchor}")

    # ==================================================================
    # ROW 3: STACKED BAR + SCATTER -- 2+ metrics, cross-filtered
    # ==================================================================
    row3_anchor = row2_anchor + 18
    if group_col and cat_start_row and len(numeric_cols) >= 2:
        section_header(row3_anchor - 1, 1, 8, f"{numeric_cols[0]} + {numeric_cols[1]} Composition % (cross-filtered)")
        stacked = BarChart()
        stacked.type = "col"
        stacked.grouping = "percentStacked"
        stacked.overlap = 100
        stacked.visible_cells_only = False
        stacked_data_ref = Reference(ws, min_col=CAT_COL + 1, max_col=CAT_COL + 2, min_row=cat_start_row, max_row=cat_end_row)
        stacked.add_data(stacked_data_ref, titles_from_data=True)
        stacked.set_categories(cats_ref)
        stacked.y_axis.numFmt = "0%"
        stacked.width, stacked.height = 15, 8
        if len(chart_groups) <= 8:
            stacked.dataLabels = DataLabelList()
            clean_labels(stacked.dataLabels, showPercent=True)
        ws.add_chart(stacked, f"A{row3_anchor}")

        section_header(row3_anchor - 1, 10, 17, f"{numeric_cols[0]} vs {numeric_cols[1]} Correlation")
        scatter = ScatterChart()
        scatter.visible_cells_only = False
        scatter.style = 13
        xvalues = Reference(ws, min_col=CAT_COL + 1, min_row=cat_start_row + 1, max_row=cat_end_row)
        yvalues = Reference(ws, min_col=CAT_COL + 2, min_row=cat_start_row, max_row=cat_end_row)
        series = Series(yvalues, xvalues, title_from_data=True)
        series.marker.symbol = "circle"
        series.marker.size = 8
        series.graphicalProperties.line.noFill = True
        scatter.series.append(series)
        scatter.x_axis.title = numeric_cols[0]
        scatter.y_axis.title = numeric_cols[1]
        scatter.width, scatter.height = 15, 8
        ws.add_chart(scatter, f"J{row3_anchor}")

    # ==================================================================
    # ROW 4: COMBO -- Bar (metric1) + Line (Achievement %), cross-filtered
    # ==================================================================
    row4_anchor = row3_anchor + 18
    if group_col and cat_start_row and len(numeric_cols) >= 2:
        ws.cell(row=cat_start_row, column=PCT_COL, value=numeric_cols[0])
        ws.cell(row=cat_start_row, column=PCT_COL + 1, value=f"{numeric_cols[1]} %")
        for r in range(cat_start_row + 1, cat_end_row + 1):
            v1 = get_column_letter(CAT_COL + 1)
            v2 = get_column_letter(CAT_COL + 2)
            ws.cell(row=r, column=PCT_COL, value=f"={v1}{r}")
            ws.cell(row=r, column=PCT_COL + 1, value=f"=IF({v1}{r}=0,0,{v2}{r}/{v1}{r}*100)")
        ws.column_dimensions[get_column_letter(PCT_COL)].hidden = True
        ws.column_dimensions[get_column_letter(PCT_COL + 1)].hidden = True

        section_header(row4_anchor - 1, 1, 17,
                        f"{numeric_cols[0]} vs {numeric_cols[1]} Achievement % by {group_col} (cross-filtered)")
        combo_bar = BarChart()
        combo_bar.visible_cells_only = False
        combo_bar_data = Reference(ws, min_col=PCT_COL, min_row=cat_start_row, max_row=cat_end_row)
        combo_bar.add_data(combo_bar_data, titles_from_data=True)
        combo_bar.set_categories(cats_ref)
        combo_bar.y_axis.title = numeric_cols[0]
        combo_bar.y_axis.majorGridlines = None
        show_combo_labels = len(chart_groups) <= 8
        if show_combo_labels:
            combo_bar.dataLabels = DataLabelList()
            clean_labels(combo_bar.dataLabels, showVal=True)
            combo_bar.dataLabels.numFmt = "#,##0"

        combo_line = LineChart()
        combo_line.visible_cells_only = False
        combo_line_data = Reference(ws, min_col=PCT_COL + 1, min_row=cat_start_row, max_row=cat_end_row)
        combo_line.add_data(combo_line_data, titles_from_data=True)
        for s in combo_line.series:
            s.smooth = False
        combo_line.y_axis.axId = 200
        combo_line.y_axis.title = "Achievement %"
        combo_line.y_axis.crosses = "max"
        if show_combo_labels:
            combo_line.dataLabels = DataLabelList()
            clean_labels(combo_line.dataLabels, showVal=True)
            combo_line.dataLabels.numFmt = '0"%"'

        combo_bar.y_axis.crosses = "autoZero"
        combo_bar += combo_line
        combo_bar.width, combo_bar.height = 30, 8
        ws.add_chart(combo_bar, f"A{row4_anchor}")

    # Pin every row's height explicitly. Charts are anchored to a cell plus a
    # FIXED physical size (cm), but their vertical position depends on how
    # tall every row above them is. Excel's default row height and Google
    # Sheets' default row height are not quite the same, so leaving row
    # height unset made the dashboard look fine on one and overlap/misalign
    # on the other. Setting an explicit height removes that ambiguity.
    last_dashboard_row = row4_anchor + 17
    for r in range(3, last_dashboard_row + 1):
        ws.row_dimensions[r].height = 15

    ws.sheet_view.showGridLines = False


def main():
    raw_folder_id = os.environ.get("RAW_FOLDER_ID")
    output_folder_id = os.environ.get("OUTPUT_FOLDER_ID")
    if not raw_folder_id or not output_folder_id:
        raise RuntimeError("RAW_FOLDER_ID and OUTPUT_FOLDER_ID environment variables must be set.")
    if raw_folder_id == output_folder_id:
        raise RuntimeError(
            "RAW_FOLDER_ID and OUTPUT_FOLDER_ID are set to the SAME folder ID -- "
            "these must be two different Drive folders. Check your GitHub Secrets."
        )
    log.info(f"Raw folder ID:    {raw_folder_id}")
    log.info(f"Output folder ID: {output_folder_id}")

    with tempfile.TemporaryDirectory() as tmp_dir:
        log.info("Connecting to Google Drive...")
        service = get_drive_service()

        log.info(f"Listing files in raw folder {raw_folder_id}...")
        files = list_data_files_in_folder(service, raw_folder_id)
        files = [f for f in files if f["name"] != REPORT_FILENAME]
        if not files:
            raise RuntimeError("No supported files (.xlsx/.xls/.csv) found in the Drive raw folder.")

        for f in files:
            dest = os.path.join(tmp_dir, f["name"])
            download_file(service, f["id"], dest)
            log.info(f"Downloaded: {f['name']}")

        log.info("Cleaning and auto-detecting structure...")
        master, numeric_cols, category_cols, date_col = build_master_dataset(tmp_dir)
        log.info(f"Total rows: {len(master)} | numeric: {numeric_cols} | "
                  f"category: {category_cols} | date_col: {date_col}")

        group_col = pick_grouping_column(category_cols, master, len(master))
        status_col = find_status_column(category_cols)
        log.info(f"Grouping summary by: {group_col} | Status column: {status_col}")

        log.info("Building formatted Excel report...")
        wb = Workbook()
        last_raw_row, col_index = write_raw_sheet(wb, master, date_col)
        write_summary_sheet(wb, master, last_raw_row, col_index, numeric_cols, group_col)
        write_dashboard_sheet(wb, master, last_raw_row, col_index, numeric_cols, group_col, date_col,
                               category_cols, status_col)

        report_path = os.path.join(tmp_dir, REPORT_FILENAME)
        wb.save(report_path)

        log.info("Uploading report back to Drive output folder...")
        file_id = upload_or_replace_file(service, report_path, output_folder_id, REPORT_FILENAME)
        log.info(f"Done. Report file ID: {file_id}")


if __name__ == "__main__":
    main()
